"""
views.py — Fase 7
Dashboard analítico + exportação de relatórios em PDF e Excel.
"""
import io
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from .services import (
    get_dados_contratos_ativos,
    get_dados_dashboard,
    get_dados_extrato,
    get_dados_inadimplencia,
    get_dados_imoveis,
)


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    dados = get_dados_dashboard()
    return render(request, 'relatorios/dashboard.html', dados)


# ─── Relatórios (listagem + filtros) ─────────────────────────────────────────

@login_required
def relatorios_index(request):
    return render(request, 'relatorios/index.html')


# ─── Extrato financeiro ───────────────────────────────────────────────────────

@login_required
def extrato(request):
    hoje = date.today()
    data_inicio = _parse_date(request.GET.get('data_inicio'), hoje.replace(day=1))
    data_fim    = _parse_date(request.GET.get('data_fim'), hoje)
    formato     = request.GET.get('formato', 'html')

    dados = get_dados_extrato(data_inicio, data_fim)

    if formato == 'pdf':
        return _render_pdf('relatorios/pdf/extrato.html', dados, 'extrato_financeiro.pdf')
    if formato == 'excel':
        return _extrato_excel(dados)

    return render(request, 'relatorios/extrato.html', dados)


# ─── Inadimplência ────────────────────────────────────────────────────────────

@login_required
def inadimplencia(request):
    formato = request.GET.get('formato', 'html')
    dados   = get_dados_inadimplencia()

    if formato == 'pdf':
        return _render_pdf('relatorios/pdf/inadimplencia.html', dados, 'inadimplencia.pdf')
    if formato == 'excel':
        return _inadimplencia_excel(dados)

    return render(request, 'relatorios/inadimplencia.html', dados)


# ─── Imóveis ─────────────────────────────────────────────────────────────────

@login_required
def imoveis(request):
    formato = request.GET.get('formato', 'html')
    dados   = get_dados_imoveis()

    if formato == 'pdf':
        return _render_pdf('relatorios/pdf/imoveis.html', dados, 'relatorio_imoveis.pdf')
    if formato == 'excel':
        return _imoveis_excel(dados)

    return render(request, 'relatorios/imoveis.html', dados)


# ─── Contratos ativos ─────────────────────────────────────────────────────────

@login_required
def contratos_ativos(request):
    formato = request.GET.get('formato', 'html')
    dados   = get_dados_contratos_ativos()

    if formato == 'pdf':
        return _render_pdf('relatorios/pdf/contratos.html', dados, 'contratos_ativos.pdf')
    if formato == 'excel':
        return _contratos_excel(dados)

    return render(request, 'relatorios/contratos.html', dados)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _parse_date(value: str | None, default: date) -> date:
    if not value:
        return default
    try:
        return date.fromisoformat(value)
    except ValueError:
        return default


def _render_pdf(template_name: str, context: dict, filename: str) -> HttpResponse:
    """Renderiza template HTML como PDF usando xhtml2pdf."""
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa

    html = render_to_string(template_name, context)
    buffer = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _extrato_excel(dados: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Extrato Financeiro'

    # Cabeçalho
    ws.merge_cells('A1:E1')
    ws['A1'] = f'Extrato Financeiro — {dados["data_inicio"].strftime("%d/%m/%Y")} a {dados["data_fim"].strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a56db')
        cell.alignment = Alignment(horizontal='center')

    for row, l in enumerate(dados['lancamentos'], 4):
        ws.cell(row=row, column=1, value=l.data_vencimento.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=l.descricao)
        ws.cell(row=row, column=3, value=l.categoria)
        ws.cell(row=row, column=4, value=l.get_tipo_display())
        ws.cell(row=row, column=5, value=float(l.valor))

    ultimo_row = ws.max_row + 2
    ws.cell(row=ultimo_row,     column=4, value='Total Receitas:').font = Font(bold=True)
    ws.cell(row=ultimo_row,     column=5, value=float(dados['total_receitas']))
    ws.cell(row=ultimo_row + 1, column=4, value='Total Despesas:').font = Font(bold=True)
    ws.cell(row=ultimo_row + 1, column=5, value=float(dados['total_despesas']))
    ws.cell(row=ultimo_row + 2, column=4, value='Saldo:').font = Font(bold=True)
    ws.cell(row=ultimo_row + 2, column=5, value=float(dados['saldo']))

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    return _excel_response(wb, 'extrato_financeiro.xlsx')


def _inadimplencia_excel(dados: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inadimplência'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Relatório de Inadimplência'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Inquilino', 'Imóvel', 'Competência', 'Vencimento', 'Dias Atraso', 'Valor (R$)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='e02424')
        cell.alignment = Alignment(horizontal='center')

    hoje = date.today()
    for row, p in enumerate(dados['parcelas'], 4):
        dias = (hoje - p.data_vencimento).days
        ws.cell(row=row, column=1, value=p.contrato.inquilino.nome)
        ws.cell(row=row, column=2, value=str(p.contrato.imovel))
        ws.cell(row=row, column=3, value=str(p.competencia))
        ws.cell(row=row, column=4, value=p.data_vencimento.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=dias)
        ws.cell(row=row, column=6, value=float(p.valor_total))

    for col in [1, 2, 3, 4]:
        ws.column_dimensions[chr(64 + col)].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16

    return _excel_response(wb, 'inadimplencia.xlsx')


def _imoveis_excel(dados: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Imóveis'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Relatório de Imóveis'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Código', 'Tipo', 'Finalidade', 'Endereço', 'Cidade', 'Status', 'Valor Aluguel']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0e9f6e')
        cell.alignment = Alignment(horizontal='center')

    for row, i in enumerate(dados['imoveis'], 4):
        ws.cell(row=row, column=1, value=i.codigo)
        ws.cell(row=row, column=2, value=i.get_tipo_display())
        ws.cell(row=row, column=3, value=i.get_finalidade_display())
        ws.cell(row=row, column=4, value=f'{i.logradouro}, {i.numero}')
        ws.cell(row=row, column=5, value=i.cidade)
        ws.cell(row=row, column=6, value=i.get_status_display())
        ws.cell(row=row, column=7, value=float(i.valor_aluguel or 0))

    widths = [10, 15, 14, 35, 18, 14, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    return _excel_response(wb, 'relatorio_imoveis.xlsx')


def _contratos_excel(dados: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contratos Ativos'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Relatório de Contratos Ativos'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Inquilino', 'Imóvel', 'Início', 'Fim', 'Aluguel (R$)', 'Reajuste', 'Garantia']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='7e3af2')
        cell.alignment = Alignment(horizontal='center')

    for row, c in enumerate(dados['contratos'], 4):
        ws.cell(row=row, column=1, value=c.inquilino.nome)
        ws.cell(row=row, column=2, value=str(c.imovel))
        ws.cell(row=row, column=3, value=c.data_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=c.data_fim.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=float(c.valor_aluguel))
        ws.cell(row=row, column=6, value=c.get_tipo_reajuste_display())
        ws.cell(row=row, column=7, value=c.get_tipo_garantia_display())

    widths = [28, 28, 12, 12, 16, 14, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    return _excel_response(wb, 'contratos_ativos.xlsx')


def _excel_response(wb, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
